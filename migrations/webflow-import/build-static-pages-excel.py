#!/usr/bin/env python3
"""
Build the Webflow global-settings + static-page SEO Excel deliverable.

Reads:
  migrations/webflow-export/site-settings.json  (GET /sites/:id)
  migrations/webflow-export/raw/pages.jsonl      (GET /sites/:id/pages, all 69)

Writes migrations/webflow-export/webflow-static-pages-seo.xlsx with:
  - "Site Settings" sheet: global config + what the Data API does NOT expose.
  - "Static Pages" sheet: live, real pages only (drafts, detail_* CMS templates,
    and *-copy / untitled junk excluded). Blank meta rendered as '-'.

Run: python3 migrations/webflow-import/build-static-pages-excel.py
"""
from __future__ import annotations

import json
import os

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

BASE = os.path.join(os.path.dirname(__file__), "..", "webflow-export")
SETTINGS = os.path.join(BASE, "site-settings.json")
PAGES = os.path.join(BASE, "raw", "pages.jsonl")
OUT = os.path.join(BASE, "webflow-static-pages-seo.xlsx")

DASH = "-"
FONT = "Arial"
HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(name=FONT, bold=True, color="FFFFFF", size=11)
CELL = Font(name=FONT, size=10)
DASHF = Font(name=FONT, size=10, color="B0B0B0")
TOP = Alignment(horizontal="left", vertical="top", wrap_text=True)
MID = Alignment(horizontal="left", vertical="center")
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def is_junk(p: dict) -> bool:
    s = (p.get("slug") or "")
    if p.get("draft") or p.get("archived"):
        return True
    if s.startswith("detail_"):
        return True
    if "copy" in s.lower() or s.startswith("untitled"):
        return True
    return False


def v(x: object) -> str:
    x = ("" if x is None else str(x)).strip()
    return x if x else DASH


def style_header(ws, headers: list[str]) -> None:
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = MID
        cell.border = BORDER


def build_settings_sheet(ws, s: dict) -> None:
    t = ws.cell(row=1, column=1, value="Webflow Global Site Settings")
    t.font = Font(name=FONT, bold=True, size=14)
    sub = ws.cell(row=2, column=1,
                  value="Source: Webflow Data API v2 GET /sites/:id (2026-06-05).")
    sub.font = Font(name=FONT, italic=True, size=9, color="595959")

    domains = ", ".join(d.get("url", "") for d in s.get("customDomains", []))
    loc = s.get("locales", {})
    primary = (loc.get("primary") or {}).get("displayName")
    secondary = loc.get("secondary") or []
    rows = [
        ("Setting", "Value"),
        ("Site name", s.get("displayName")),
        ("Custom domains", domains or DASH),
        ("Primary locale", primary or DASH),
        ("Secondary locales", ", ".join(x.get("displayName", "") for x in secondary) or "none (single-locale → no hreflang)"),
        ("Time zone", s.get("timeZone")),
        ("Cookie/data collection", "enabled" if s.get("dataCollectionEnabled") else "disabled"),
        ("Google Tag IDs", ", ".join(s.get("googleTagIds") or []) or "none set in Webflow"),
        ("Last published", s.get("lastPublished")),
        ("Last updated", s.get("lastUpdated")),
    ]
    start = 4
    for i, (k, val) in enumerate(rows):
        r = start + i
        kc = ws.cell(row=r, column=1, value=k)
        vc = ws.cell(row=r, column=2, value=v(val))
        if i == 0:
            kc.fill = vc.fill = HEADER_FILL
            kc.font = vc.font = HEADER_FONT
        else:
            kc.font = Font(name=FONT, bold=True, size=10)
            vc.font = CELL
        kc.alignment = vc.alignment = MID
        kc.border = vc.border = BORDER

    note_r = start + len(rows) + 1
    nh = ws.cell(row=note_r, column=1, value="NOT available via Webflow Data API")
    nh.font = Font(name=FONT, bold=True, size=11, color="C00000")
    notes = [
        "301 redirects — not exposed in the API (were configured in Site Settings → Publishing).",
        "robots.txt — not exposed in the API.",
        "Custom head/footer code — requires custom_code scope (not granted on this token).",
        "Global SEO defaults (minify, indexing toggles) — not exposed; only per-page SEO is.",
    ]
    for i, n in enumerate(notes):
        c = ws.cell(row=note_r + 1 + i, column=1, value=f"•  {n}")
        c.font = Font(name=FONT, size=10, color="595959")

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 70
    ws.sheet_view.showGridLines = False


def build_pages_sheet(ws, pages: list[dict]) -> None:
    headers = ["Page Name", "Slug", "Webflow URL", "Meta Title",
               "Meta Description", "OG Title", "OG Description", "Last Updated"]
    style_header(ws, headers)
    kept = sorted((p for p in pages if not is_junk(p)),
                  key=lambda x: (x.get("publishedPath") or ""))
    for p in kept:
        seo = p.get("seo") or {}
        og = p.get("openGraph") or {}
        ws.append([
            v(p.get("title")),
            v(p.get("slug")) if p.get("slug") else "(home/root)",
            v(p.get("publishedPath")),
            v(seo.get("title")),
            v(seo.get("description")),
            v(og.get("title")),
            v(og.get("description")),
            v((p.get("lastUpdated") or "")[:10]),
        ])
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.font = DASHF if cell.value == DASH else CELL
            cell.alignment = TOP
            cell.border = BORDER
    for i, w in enumerate([30, 32, 30, 50, 70, 30, 50, 14], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False


def main() -> None:
    settings = json.load(open(SETTINGS, encoding="utf-8"))
    pages = [json.loads(l) for l in open(PAGES, encoding="utf-8") if l.strip()]
    wb = Workbook()
    build_settings_sheet(wb.active, settings)
    wb.active.title = "Site Settings"
    build_pages_sheet(wb.create_sheet("Static Pages"), pages)
    wb.save(OUT)
    kept = [p for p in pages if not is_junk(p)]
    print(f"[xlsx] wrote {OUT}  ({len(kept)} live pages)")


if __name__ == "__main__":
    main()
