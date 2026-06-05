#!/usr/bin/env python3
"""
Build the single combined Webflow SEO workbook.

Sheets:
  1. Summary       — global site settings, per-source stats, API gaps.
  2. Static Pages  — 40 live real Webflow pages (drafts / detail_* / junk excluded).
  3..14. One sheet per CMS collection (Blogs … About Galleries).

Sources:
  migrations/webflow-export/seo-extract.csv     (CMS collection meta)
  migrations/webflow-export/site-settings.json  (GET /sites/:id)
  migrations/webflow-export/raw/pages.jsonl      (GET /sites/:id/pages)

Run: python3 migrations/webflow-import/build-combined-excel.py
"""
from __future__ import annotations

import csv
import json
import os
from collections import OrderedDict

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

BASE = os.path.join(os.path.dirname(__file__), "..", "webflow-export")
SRC = os.path.join(BASE, "seo-extract.csv")
SETTINGS = os.path.join(BASE, "site-settings.json")
PAGES = os.path.join(BASE, "raw", "pages.jsonl")
REDIRECTS = os.path.join(BASE, "redirects-final.json")
CODE_SEO = os.path.join(BASE, "global-code-seo.json")
HEAD_AUDIT = os.path.join(BASE, "head-audit.json")
BROKEN = os.path.join(BASE, "broken-links.json")
PARITY = os.path.join(BASE, "url-parity.json")
SITEMAP_PARITY = os.path.join(BASE, "sitemap-parity.json")
OG_COVERAGE = os.path.join(BASE, "og-coverage.json")
LEGAL = os.path.join(BASE, "legal-pages.json")
NEW_DESC = os.path.join(BASE, "new-meta-descriptions.csv")
OUT = os.path.join(BASE, "webflow-seo-export.xlsx")


def load_new_descriptions() -> dict:
    out: dict = {}
    if os.path.exists(NEW_DESC):
        with open(NEW_DESC, encoding="utf-8") as fh:
            for row in csv.DictReader(fh):
                out[(row["collection"], row["slug"])] = row["new_meta_description"]
    return out

# Recommended action for each 404-risk old URL.
PARITY_REC = {
    "/pricing": "HIGH: build /pricing or redirect → /book-a-demo",
    "/acceptable-use-policy": "redirect → /legal/acceptable-use-policy",
    "/leadership": "redirect → /teams",
    "/webinar": "redirect → /webinars",
    "/search": "utility — redirect → / or rebuild search",
    "/survey": "utility — redirect → / or rebuild",
    "/about-copy": "junk copy — 410 Gone",
    "/pricing-copy": "junk copy — 410 Gone",
    "/cleanstart-hitachi-bengaluru": "decide: rebuild or redirect/410",
    "/cleanstart-hitachi-chennai": "decide: rebuild or redirect/410",
    "/cleanstart-hitachi-hyderabad": "decide: rebuild or redirect/410",
    "/cleanstart-raksha-chennai": "decide: rebuild or redirect/410",
    "/new-year-event-eventus": "redirect/410 (one-off event)",
    "/new-year-event-sysdig": "redirect/410 (one-off event)",
}

# Known correct targets for the broken guide-body links (real guide slugs exist).
BROKEN_FIX = {
    "/cdn-cgi/l/email-protection": "— (Cloudflare email-obfuscation artifact; not a real page, gone on new site)",
    "/container": "/guide/container",
    "/container/image": "/guide/container-image",
    "/container/image/dockerimage": "/guide/docker-images",
    "/container/image/buildsi-cd-pipeline": "/guide/ci-cd (review)",
    "/container/sbom": "/guide/sbom",
    "/guide/cgroup%5C": "/guide/cgroup (stray backslash in href)",
    "/guide/orchestration": "/guide/container-orchestration",
    "/guide/security": "/guide/container-security",
}

COLLECTIONS: "OrderedDict[str, tuple[str, str | None]]" = OrderedDict([
    ("blogs", ("Blogs", "/blogs")),
    ("news", ("News", "/news")),
    ("guides", ("Guides", "/guide")),
    ("events", ("Events", "/event")),
    ("jobs", ("Jobs", "/job")),
    ("authors", ("Authors", "/author")),
    ("resources", ("Resources", "/resources")),
    ("webinars", ("Webinars", "/webinar")),
    ("categories", ("Categories", "/category")),
    ("newsCategories", ("News Categories", "/news-categories")),
    ("jobLocations", ("Job Locations", None)),
    ("aboutGalleries", ("About Galleries", None)),
])

DASH = "-"
FONT = "Arial"
HFILL = PatternFill("solid", fgColor="1F3864")
HFONT = Font(name=FONT, bold=True, color="FFFFFF", size=11)
CELL = Font(name=FONT, size=10)
DASHF = Font(name=FONT, size=10, color="B0B0B0")
BOLD = Font(name=FONT, bold=True, size=10)
TOP = Alignment(horizontal="left", vertical="top", wrap_text=True)
MID = Alignment(horizontal="left", vertical="center")
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def v(x: object) -> str:
    x = ("" if x is None else str(x)).strip()
    return x if x else DASH


def load_collections() -> "OrderedDict[str, list[dict]]":
    data: "OrderedDict[str, list[dict]]" = OrderedDict((s, []) for s in COLLECTIONS)
    with open(SRC, encoding="utf-8") as fh:
        for row in csv.DictReader(fh):
            if row["collection"] in data:
                data[row["collection"]].append(row)
    for rows in data.values():
        rows.sort(key=lambda r: r["slug"])
    return data


def page_is_junk(p: dict) -> bool:
    s = p.get("slug") or ""
    if p.get("draft") or p.get("archived"):
        return True
    return s.startswith("detail_") or "copy" in s.lower() or s.startswith("untitled")


def header_row(ws, headers: list[str], row: int = 1) -> None:
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.fill = HFILL
        cell.font = HFONT
        cell.alignment = MID
        cell.border = BORDER


def body_style(ws, ncol: int, start: int = 2) -> None:
    for row in ws.iter_rows(min_row=start, max_row=ws.max_row, min_col=1, max_col=ncol):
        for cell in row:
            cell.font = DASHF if cell.value == DASH else CELL
            cell.alignment = TOP
            cell.border = BORDER


# ── Sheet 1: Summary ────────────────────────────────────────────────
def build_summary(ws, settings, data, live_pages, redirects, audits) -> None:
    ws.cell(row=1, column=1, value="Webflow → CleanStart SEO Export").font = Font(
        name=FONT, bold=True, size=14)
    ws.cell(row=2, column=1,
            value="Source: Webflow Data API v2 (2026-06-05). CMS-collection meta "
                  "imported verbatim into local + prod CMS. '-' = null/blank.").font = Font(
        name=FONT, italic=True, size=9, color="595959")

    # Global settings block
    domains = ", ".join(d.get("url", "") for d in settings.get("customDomains", []))
    loc = settings.get("locales", {})
    secondary = loc.get("secondary") or []
    gset = [
        ("Site name", settings.get("displayName")),
        ("Custom domains", domains or DASH),
        ("Primary locale", (loc.get("primary") or {}).get("displayName") or DASH),
        ("Secondary locales", ", ".join(x.get("displayName", "") for x in secondary)
         or "none (single-locale → no hreflang)"),
        ("Time zone", settings.get("timeZone")),
        ("Cookie/data collection", "enabled" if settings.get("dataCollectionEnabled") else "disabled"),
        ("Google Tag IDs", ", ".join(settings.get("googleTagIds") or []) or "none set in Webflow"),
        ("Last published", settings.get("lastPublished")),
    ]
    r = 4
    ws.cell(row=r, column=1, value="GLOBAL SITE SETTINGS").font = Font(
        name=FONT, bold=True, size=11)
    r += 1
    header_row(ws, ["Setting", "Value"], r)
    r += 1
    for k, val in gset:
        ws.cell(row=r, column=1, value=k).font = BOLD
        ws.cell(row=r, column=2, value=v(val)).font = CELL
        for c in (1, 2):
            ws.cell(row=r, column=c).border = BORDER
            ws.cell(row=r, column=c).alignment = MID
        r += 1

    # Stats table
    r += 1
    ws.cell(row=r, column=1, value="CONTENT STATS").font = Font(name=FONT, bold=True, size=11)
    r += 1
    header_row(ws, ["Source", "Total Items", "Has Meta Title", "Has Meta Description"], r)
    r += 1
    stat_start = r

    def stat(rows, tkey, dkey):
        return (len(rows),
                sum(1 for x in rows if (x.get(tkey) or "").strip()),
                sum(1 for x in rows if (x.get(dkey) or "").strip()))

    # Static pages first
    lp = [{"t": (p.get("seo") or {}).get("title"),
           "d": (p.get("seo") or {}).get("description")} for p in live_pages]
    n, t, d = stat(lp, "t", "d")
    ws.cell(row=r, column=1, value="Static Pages (live)").font = CELL
    ws.cell(row=r, column=2, value=n).font = CELL
    ws.cell(row=r, column=3, value=t).font = CELL
    ws.cell(row=r, column=4, value=d).font = CELL
    for c in range(1, 5):
        ws.cell(row=r, column=c).border = BORDER
        ws.cell(row=r, column=c).alignment = MID
    r += 1

    for stem, (name, _) in COLLECTIONS.items():
        n, t, d = stat(data[stem], "meta_title", "meta_description")
        ws.cell(row=r, column=1, value=name).font = CELL
        ws.cell(row=r, column=2, value=n).font = CELL
        ws.cell(row=r, column=3, value=t).font = CELL
        ws.cell(row=r, column=4, value=d).font = CELL
        for c in range(1, 5):
            ws.cell(row=r, column=c).border = BORDER
            ws.cell(row=r, column=c).alignment = MID
        r += 1

    # TOTAL (hardcoded — no LibreOffice available to recalc formulas here)
    rows_for_total = list(range(stat_start, r))
    tot = [sum(ws.cell(row=rr, column=c).value or 0 for rr in rows_for_total)
           for c in (2, 3, 4)]
    ws.cell(row=r, column=1, value="TOTAL").font = BOLD
    ws.cell(row=r, column=1).border = BORDER
    for c, val in zip((2, 3, 4), tot):
        cell = ws.cell(row=r, column=c, value=val)
        cell.font = BOLD
        cell.border = BORDER
    r += 2

    # Redirect stats
    rlist = redirects.get("redirects", [])
    rstruct = redirects.get("structural", [])
    rchains = sum(1 for e in rlist if e.get("chain"))
    ws.cell(row=r, column=1, value="REDIRECTS").font = Font(name=FONT, bold=True, size=11)
    r += 1
    header_row(ws, ["Source", "Count"], r)
    r += 1
    rrows = [
        ("Webflow content 301s", len(rlist)),
        ("…of which multi-hop chains (collapsed)", rchains),
        ("Structural (apex→www, HTTP→HTTPS)", len(rstruct)),
        ("Cloudflare custom redirects (rules/page-rules/bulk)", 0),
        ("TOTAL redirects", len(rlist) + len(rstruct)),
    ]
    for i, (k, val) in enumerate(rrows):
        bold = (i == len(rrows) - 1)
        ws.cell(row=r, column=1, value=k).font = BOLD if bold else CELL
        ws.cell(row=r, column=2, value=val).font = BOLD if bold else CELL
        for c in (1, 2):
            ws.cell(row=r, column=c).border = BORDER
            ws.cell(row=r, column=c).alignment = MID
        r += 1
    ws.cell(row=r, column=1, value="↳ See the Redirects sheet for the full source→destination list.").font = Font(
        name=FONT, italic=True, size=9, color="595959")
    r += 2

    # API gaps — all now captured via the admin UI
    ws.cell(row=r, column=1, value="API-gap items — captured via Webflow admin UI (not in Data API)").font = Font(
        name=FONT, bold=True, size=11, color="2E7D32")
    r += 1
    for n in ("301 redirects — API is Enterprise-only; 26 rules exported from admin UI. ✓ see Redirects sheet",
              "Global custom code (GA4, AdSense, LinkedIn, HubSpot, Leadfeeder) — ✓ see Global Code & SEO sheet",
              "robots.txt + global SEO settings (AI bots, canonical, sitemap, llms.txt) — ✓ see Global Code & SEO sheet",
              "Per-page SEO defaults — captured per page (Static Pages sheet) + per CMS item (collection sheets)."):
        ws.cell(row=r, column=1, value=f"•  {n}").font = Font(
            name=FONT, size=10, color="595959")
        r += 1

    # Key go-live findings
    url404 = audits.get("parity", {}).get("counts", {}).get("404 RISK", 0)
    dropped = audits.get("sitemap", {}).get("counts", {}).get("old_only", 0)
    broken_real = max(0, audits.get("broken", {}).get("counts", {}).get("broken", 0) - 1)
    RED, AMBER, GREEN = "C00000", "B8860B", "2E7D32"
    findings = [
        (f"Old indexed URLs that would 404 at go-live (incl /pricing)", url404, "HIGH", "URL Parity", RED),
        ("Authors are DRAFT → /author pages 404 + dropped from sitemap", 6, "HIGH", "Sitemap Parity", RED),
        ("News post slug truncated in CMS → old indexed URL 404s", 1, "HIGH", "Sitemap Parity", RED),
        ("URLs dropped from sitemap (advertised today, not in new sitemap)", dropped, "MED", "Sitemap Parity", AMBER),
        ("Broken internal links inside imported guide bodies", broken_real, "MED", "Broken Links", AMBER),
        ("GA4 (G-S6T47D7PZR) lives only in Webflow custom code — wire into new site", "—", "MED", "Global Code & SEO", AMBER),
        (".webflow.io subdomain indexing ON — block before launch", "—", "MED", "Global Code & SEO", AMBER),
        ("Old-site canonical bugs (news wrong / guide+event missing) — FIXED by new site", 3, "INFO", "Canonical & Head Audit", GREEN),
        ("hreflang: none site-wide (single-locale) — nothing to migrate", "—", "INFO", "Canonical & Head Audit", GREEN),
        ("OG images: new site 100% coverage via dynamic fallback — improvement", "—", "GOOD", "OG Image Coverage", GREEN),
    ]
    r += 1
    ws.cell(row=r, column=1, value="KEY GO-LIVE FINDINGS (see linked sheets)").font = Font(
        name=FONT, bold=True, size=11)
    r += 1
    header_row(ws, ["Finding", "Count", "Severity", "Sheet"], r)
    r += 1
    for text, cnt, sev, sheet, color in findings:
        ws.cell(row=r, column=1, value=text).font = CELL
        ws.cell(row=r, column=2, value=cnt).font = CELL
        sc = ws.cell(row=r, column=3, value=sev)
        sc.font = Font(name=FONT, bold=True, size=10, color=color)
        ws.cell(row=r, column=4, value=sheet).font = CELL
        for c in range(1, 5):
            ws.cell(row=r, column=c).border = BORDER
            ws.cell(row=r, column=c).alignment = MID
        r += 1

    # Workbook contents
    r += 1
    ws.cell(row=r, column=1, value="WORKBOOK CONTENTS").font = Font(name=FONT, bold=True, size=11)
    r += 1
    contents = [
        "Summary — this dashboard: settings, stats, redirect counts, key findings",
        "Redirects — 26 authoritative Webflow 301s + structural (chains resolved)",
        "URL Parity — every old URL → exists / redirected / 404 on new site",
        "Sitemap Parity — old sitemap vs new sitemap.ts (dropped / added)",
        "Global Code & SEO — GA4/HubSpot/LinkedIn etc. + robots/SEO settings",
        "Canonical & Head Audit — canonical / hreflang / OG / JSON-LD per page type",
        "OG Image Coverage — social-share image source per content type",
        "Broken Links — broken internal links found by crawling the live site",
        "Static Pages — 40 live static pages: slug, URL, meta, OG",
        "Blogs … About Galleries — per-collection meta (slug, URL, title, description)",
    ]
    for c in contents:
        ws.cell(row=r, column=1, value=f"•  {c}").font = Font(name=FONT, size=10, color="595959")
        r += 1

    for col, w in zip("ABCD", (62, 10, 12, 24)):
        ws.column_dimensions[col].width = w
    ws.sheet_view.showGridLines = False


# ── Sheet 2: Static Pages ───────────────────────────────────────────
def build_static_pages(ws, live_pages) -> None:
    headers = ["Page Name", "Slug", "Webflow URL", "Meta Title",
               "Meta Description", "OG Title", "OG Description", "Last Updated"]
    header_row(ws, headers)
    for p in sorted(live_pages, key=lambda x: (x.get("publishedPath") or "")):
        seo = p.get("seo") or {}
        og = p.get("openGraph") or {}
        ws.append([
            v(p.get("title")),
            v(p.get("slug")) if p.get("slug") else "(home/root)",
            v(p.get("publishedPath")),
            v(seo.get("title")), v(seo.get("description")),
            v(og.get("title")), v(og.get("description")),
            v((p.get("lastUpdated") or "")[:10]),
        ])
    body_style(ws, len(headers))
    for i, w in enumerate([30, 32, 28, 50, 68, 28, 48, 14], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False


# ── Legal sheet ─────────────────────────────────────────────────────
def build_legal(ws, lg: dict) -> None:
    ws.cell(row=1, column=1, value="Legal Pages (live site)").font = Font(
        name=FONT, bold=True, size=14)
    ws.cell(row=2, column=1, value=lg.get("_note", "")).font = Font(
        name=FONT, italic=True, size=9, color="595959")
    headers = ["Page Name", "Slug", "URL", "Status", "Date shown on page",
               "Created", "Last Updated", "Meta Title", "Meta Description"]
    hr = 4
    header_row(ws, headers, hr)
    r = hr + 1
    for row in lg.get("rows", []):
        vals = [row.get("name"), row.get("slug"), row.get("url"), row.get("status"),
                row.get("onpage_date"), row.get("created"), row.get("last_updated"),
                row.get("meta_title"), row.get("meta_description")]
        bad = "UNFILLED" in (row.get("onpage_date") or "") or "none" in (row.get("onpage_date") or "")
        for c, val in enumerate(vals, start=1):
            cell = ws.cell(row=r, column=c, value=v(val))
            cell.alignment = TOP
            cell.border = BORDER
            if c == 5 and bad:
                cell.font = Font(name=FONT, size=10, color="C00000")
            else:
                cell.font = DASHF if cell.value == DASH else CELL
        r += 1
    r += 1
    ws.cell(row=r, column=1, value="Documents bundled inside /legal: " +
            ", ".join(lg.get("bundled_in_legal_hub", []))).font = Font(
        name=FONT, italic=True, size=9, color="595959")
    r += 2
    for c in lg.get("conclusions", []):
        ws.cell(row=r, column=1, value=f"•  {c}").font = Font(name=FONT, size=10, color="595959")
        r += 1
    for i, w in enumerate([30, 24, 26, 18, 44, 12, 14, 28, 30], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = f"A{hr + 1}"
    ws.sheet_view.showGridLines = False


# ── Collection sheets ───────────────────────────────────────────────
def build_collection(ws, stem, rows, new_desc) -> None:
    _, prefix = COLLECTIONS[stem]
    headers = ["Name", "Slug", "URL", "Meta Title", "Meta Description (Webflow)",
               "Updated Meta Description", "Created", "Last Updated", "Last Published"]
    header_row(ws, headers)
    for rrow in rows:
        slug = rrow["slug"].strip()
        url = f"{prefix}/{slug}" if (prefix and slug) else DASH
        ws.append([v(rrow["name"]), v(slug), url,
                   v(rrow["meta_title"]), v(rrow["meta_description"]),
                   v(new_desc.get((stem, slug))),
                   v(rrow.get("created")), v(rrow.get("last_updated")),
                   v(rrow.get("last_published"))])
    body_style(ws, len(headers))
    for i, w in enumerate([34, 44, 46, 58, 70, 70, 12, 14, 14], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False


# ── Redirects sheet ─────────────────────────────────────────────────
def build_redirects(ws, rd: dict) -> None:
    ws.cell(row=1, column=1, value="Old-Site Redirects (authoritative)").font = Font(
        name=FONT, bold=True, size=14)
    ws.cell(row=2, column=1, value=(
        f"Source: {rd.get('authoritative_source')} ({rd.get('checked_at')}). "
        f"{rd.get('webflow_count')} Webflow content 301s + "
        f"{len(rd.get('structural', []))} structural. 'Final Destination' collapses "
        "multi-hop chains — seed the NEW site with the FINAL destination, not the "
        "intermediate hop.")).font = Font(name=FONT, italic=True, size=9, color="595959")
    cf = rd.get("cloudflare_layer", "")
    ws.cell(row=3, column=1, value=(
        f"Cloudflare layer: {cf}. The structural apex→www / HTTPS redirects below "
        "are host-level and already reimplemented in apps/web (proxy.ts).")).font = Font(
        name=FONT, italic=True, size=9, color="C00000")

    headers = ["Source (old URL)", "Configured Target", "Final Destination",
               "Hops", "Status", "Layer", "Live-verified"]
    hr = 5
    header_row(ws, headers, hr)
    r = hr + 1
    allrows = list(rd.get("redirects", [])) + list(rd.get("structural", []))
    for x in allrows:
        ws.cell(row=r, column=1, value=v(x.get("source")))
        ws.cell(row=r, column=2, value=v(x.get("configured_target")))
        ws.cell(row=r, column=3, value=v(x.get("final_destination")))
        ws.cell(row=r, column=4, value=v(x.get("hops")))
        ws.cell(row=r, column=5, value=v(x.get("status")))
        ws.cell(row=r, column=6, value=v(x.get("layer")))
        ws.cell(row=r, column=7, value="yes" if x.get("live_verified") else "no (rule exists, source not linked)")
        r += 1
    for row in ws.iter_rows(min_row=hr + 1, max_row=ws.max_row, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.font = DASHF if cell.value == DASH else CELL
            cell.alignment = TOP
            cell.border = BORDER
    for i, w in enumerate([46, 34, 36, 7, 8, 22, 30], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = f"A{hr + 1}"
    ws.sheet_view.showGridLines = False


# ── Global Code & SEO sheet ─────────────────────────────────────────
def build_code_seo(ws, cs: dict) -> None:
    ws.cell(row=1, column=1, value="Global Custom Code & SEO Settings").font = Font(
        name=FONT, bold=True, size=14)
    ws.cell(row=2, column=1, value=(
        "Extracted from Webflow admin (Custom code + SEO tabs), 2026-06-05 — NOT "
        "available via the Data API. Reproduce the 'migrate: Yes' scripts on the new "
        "site for analytics/marketing continuity.")).font = Font(
        name=FONT, italic=True, size=9, color="595959")

    r = 4
    ws.cell(row=r, column=1, value="TRACKING & MARKETING SCRIPTS (global head/footer code)").font = Font(
        name=FONT, bold=True, size=11)
    r += 1
    th = ["Service", "Identifier", "Location", "Purpose", "Migrate to new site?"]
    header_row(ws, th, r)
    r += 1
    for s in cs.get("tracking_scripts", []):
        vals = [s.get("service"), s.get("identifier"), s.get("location"),
                s.get("purpose"), s.get("migrate")]
        for c, val in enumerate(vals, start=1):
            cell = ws.cell(row=r, column=c, value=v(val))
            cell.font = CELL
            cell.alignment = TOP
            cell.border = BORDER
        r += 1

    r += 1
    ws.cell(row=r, column=1, value="GLOBAL SEO SETTINGS").font = Font(
        name=FONT, bold=True, size=11)
    r += 1
    header_row(ws, ["Setting", "Value"], r)
    r += 1
    for s in cs.get("seo_settings", []):
        ws.cell(row=r, column=1, value=v(s.get("setting"))).font = BOLD
        c2 = ws.cell(row=r, column=2, value=v(s.get("value")))
        c2.font = CELL
        for c in (1, 2):
            ws.cell(row=r, column=c).border = BORDER
            ws.cell(row=r, column=c).alignment = TOP
        r += 1

    for i, w in enumerate([34, 30, 12, 40, 46], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A5"
    ws.sheet_view.showGridLines = False


# ── Canonical & Head Audit sheet ────────────────────────────────────
def build_head_audit(ws, ha: dict) -> None:
    ws.cell(row=1, column=1, value="Canonical / hreflang / Social Head-Tag Audit").font = Font(
        name=FONT, bold=True, size=14)
    ws.cell(row=2, column=1, value=ha.get("_note", "")).font = Font(
        name=FONT, italic=True, size=9, color="595959")

    headers = ["Page type", "Sample URL", "Canonical", "hreflang",
               "og:title/desc", "og:image", "robots", "JSON-LD", "Status / Notes"]
    hr = 4
    header_row(ws, headers, hr)
    r = hr + 1
    for row in ha.get("rows", []):
        vals = [row.get("page_type"), row.get("sample"), row.get("canonical"),
                row.get("hreflang"), row.get("og_title_desc"), row.get("og_image"),
                row.get("robots"), row.get("jsonld"), row.get("status")]
        bug = "BROKEN" in (row.get("canonical") or "") or "MISSING" in (row.get("canonical") or "")
        for c, val in enumerate(vals, start=1):
            cell = ws.cell(row=r, column=c, value=v(val))
            cell.alignment = TOP
            cell.border = BORDER
            # flag canonical-bug rows in red on the Canonical + Status columns
            if bug and c in (3, 9):
                cell.font = Font(name=FONT, size=10, color="C00000")
            else:
                cell.font = DASHF if cell.value == DASH else CELL
        r += 1

    r += 1
    ws.cell(row=r, column=1, value="CONCLUSIONS").font = Font(name=FONT, bold=True, size=11)
    r += 1
    for c in ha.get("conclusions", []):
        ws.cell(row=r, column=1, value=f"•  {c}").font = Font(name=FONT, size=10, color="595959")
        r += 1

    for i, w in enumerate([16, 22, 40, 10, 14, 26, 18, 10, 50], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = f"A{hr + 1}"
    ws.sheet_view.showGridLines = False


# ── Broken Links sheet ──────────────────────────────────────────────
def build_broken_links(ws, bl: dict) -> None:
    ws.cell(row=1, column=1, value="Broken Internal Links (live-crawled)").font = Font(
        name=FONT, bold=True, size=14)
    ws.cell(row=2, column=1, value=(
        f"Crawled {bl.get('sitemap_pages')} sitemap pages, checked "
        f"{bl.get('internal_targets_checked')} unique internal targets "
        f"({bl.get('checked_at')}). 0 sitemap pages broken. Most broken links live in "
        "guide body content (imported into the CMS) — fix in Payload so the new site "
        "isn't broken too.")).font = Font(name=FONT, italic=True, size=9, color="595959")

    headers = ["Broken URL", "HTTP", "# linking pages", "Linked from (sample)", "Likely correct target"]
    hr = 4
    header_row(ws, headers, hr)
    r = hr + 1
    for b in bl.get("broken", []):
        url = b.get("url")
        benign = url == "/cdn-cgi/l/email-protection"
        vals = [url, b.get("status"), b.get("link_count"),
                ", ".join(b.get("linked_from", [])[:6]), BROKEN_FIX.get(url, "review")]
        for c, val in enumerate(vals, start=1):
            cell = ws.cell(row=r, column=c, value=v(val))
            cell.alignment = TOP
            cell.border = BORDER
            if benign:
                cell.font = DASHF
            elif c == 1:
                cell.font = Font(name=FONT, size=10, color="C00000")
            else:
                cell.font = CELL
        r += 1

    redir = bl.get("redirected_internal_links", [])
    if redir:
        r += 1
        ws.cell(row=r, column=1, value="INTERNAL LINKS POINTING AT A REDIRECT (update to link directly)").font = Font(
            name=FONT, bold=True, size=10, color="B8860B")
        r += 1
        header_row(ws, ["URL linked", "HTTP", "Redirects to", "# linking pages"], r)
        r += 1
        for x in redir:
            for c, val in enumerate([x.get("url"), x.get("status"), x.get("location"), x.get("link_count")], start=1):
                cell = ws.cell(row=r, column=c, value=v(val))
                cell.font = CELL
                cell.alignment = TOP
                cell.border = BORDER
            r += 1

    for i, w in enumerate([40, 8, 14, 56, 38], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = f"A{hr + 1}"
    ws.sheet_view.showGridLines = False


# ── URL Parity sheet ────────────────────────────────────────────────
def build_url_parity(ws, pa: dict) -> None:
    c = pa.get("counts", {})
    ws.cell(row=1, column=1, value="URL Parity — old indexed URLs → new site").font = Font(
        name=FONT, bold=True, size=14)
    ws.cell(row=2, column=1, value=(
        f"Every old live URL ({pa.get('old_url_count')} from {pa.get('old_urls_from')}) "
        f"checked against new apps/web routes + CMS slugs ({pa.get('checked_at')}). "
        f"OK: {c.get('OK', 0)} · 404 RISK: {c.get('404 RISK', 0)} · CHECK: {c.get('CHECK', 0)}. "
        "404-RISK = indexed today, would break at go-live without a redirect.")).font = Font(
        name=FONT, italic=True, size=9, color="595959")

    headers = ["Old URL (indexed)", "New-site status", "Detail", "Recommended action"]
    hr = 4
    header_row(ws, headers, hr)
    r = hr + 1
    order = {"404 RISK": 0, "CHECK": 1, "OK": 2}
    for row in sorted(pa.get("rows", []), key=lambda x: (order.get(x["status"], 9), x["old_url"])):
        risk = row["status"] != "OK"
        rec = PARITY_REC.get(row["old_url"])
        if rec is None:
            if row["old_url"].startswith("/webinar/"):
                rec = "redirect → /webinars (no webinar detail route built)"
            else:
                rec = "" if not risk else "review"
        vals = [row["old_url"], row["status"], row["detail"], rec]
        for cc, val in enumerate(vals, start=1):
            cell = ws.cell(row=r, column=cc, value=v(val))
            cell.alignment = TOP
            cell.border = BORDER
            if risk and cc in (1, 2):
                cell.font = Font(name=FONT, size=10, color="C00000")
            else:
                cell.font = DASHF if cell.value == DASH else CELL
        r += 1
    for i, w in enumerate([56, 14, 46, 44], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = f"A{hr + 1}"
    ws.sheet_view.showGridLines = False


# ── Sitemap Parity sheet ────────────────────────────────────────────
def _sitemap_reason(u: str) -> tuple[str, str]:
    if u.startswith("/author/"):
        return ("Authors are DRAFT in CMS (no publishedAt)", "Publish authors, or 410/redirect — pages likely 404 on new site")
    if "sysdig-to-deliver" in u:
        return ("CMS slug truncated to '…-from'", "Redirect old→new slug, or fix the CMS slug to match")
    if u.startswith("/webinar"):
        return ("No webinar detail route on new site", "redirect → /webinars (or build route)")
    rec = PARITY_REC.get(u)
    if rec:
        return ("Not emitted by new sitemap.ts", rec)
    return ("Not in new sitemap", "review")


def build_sitemap_parity(ws, sp: dict) -> None:
    c = sp.get("counts", {})
    ws.cell(row=1, column=1, value="Sitemap Parity — old sitemap vs new sitemap.ts").font = Font(
        name=FONT, bold=True, size=14)
    ws.cell(row=2, column=1, value=(
        f"Old live sitemap: {c.get('old')} URLs · New sitemap.ts (reconstructed from "
        f"prod CMS + static routes + knowledge-hub): {c.get('new')} · matched {c.get('matched')} · "
        f"dropped {c.get('old_only')} · added {c.get('new_only')}. "
        "'Dropped' = advertised to search engines today but NOT in the new sitemap.")).font = Font(
        name=FONT, italic=True, size=9, color="595959")

    headers = ["Dropped URL (in old sitemap, not new)", "Why dropped", "Recommendation"]
    hr = 4
    header_row(ws, headers, hr)
    r = hr + 1
    for u in sp.get("old_only", []):
        why, rec = _sitemap_reason(u)
        for cc, val in enumerate([u, why, rec], start=1):
            cell = ws.cell(row=r, column=cc, value=v(val))
            cell.alignment = TOP
            cell.border = BORDER
            cell.font = Font(name=FONT, size=10, color="C00000") if cc == 1 else CELL
        r += 1

    r += 1
    ws.cell(row=r, column=1, value=f"NEW-only (added to sitemap — new content, not a problem): {c.get('new_only', 0)} URLs").font = Font(
        name=FONT, bold=True, size=10, color="2E7D32")
    r += 1
    for u in sp.get("new_only", []):
        ws.cell(row=r, column=1, value=f"•  {u}").font = Font(name=FONT, size=9, color="595959")
        r += 1

    for i, w in enumerate([62, 38, 50], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = f"A{hr + 1}"
    ws.sheet_view.showGridLines = False


# ── OG Image Coverage sheet ─────────────────────────────────────────
def build_og_coverage(ws, og: dict) -> None:
    ws.cell(row=1, column=1, value="OG Image Coverage (social share images)").font = Font(
        name=FONT, bold=True, size=14)
    ws.cell(row=2, column=1, value=og.get("_note", "")).font = Font(
        name=FONT, italic=True, size=9, color="595959")

    headers = ["Content type", "Pages", "Image source in CMS", "og:image OLD site", "og:image NEW site", "Status"]
    hr = 4
    header_row(ws, headers, hr)
    r = hr + 1
    for row in og.get("rows", []):
        vals = [row.get("type"), row.get("pages"), row.get("image_source"),
                row.get("og_old"), row.get("og_new"), row.get("status")]
        gap = "NONE" in (row.get("og_old") or "") or "improved" in (row.get("status") or "")
        for cc, val in enumerate(vals, start=1):
            cell = ws.cell(row=r, column=cc, value=v(val))
            cell.alignment = TOP
            cell.border = BORDER
            if cc == 6 and "improved" in (row.get("status") or ""):
                cell.font = Font(name=FONT, size=10, color="2E7D32")
            elif cc == 4 and "NONE" in (row.get("og_old") or ""):
                cell.font = Font(name=FONT, size=10, color="C00000")
            else:
                cell.font = CELL
        r += 1

    r += 1
    ws.cell(row=r, column=1, value="CONCLUSIONS").font = Font(name=FONT, bold=True, size=11)
    r += 1
    for c in og.get("conclusions", []):
        ws.cell(row=r, column=1, value=f"•  {c}").font = Font(name=FONT, size=10, color="595959")
        r += 1

    for i, w in enumerate([18, 8, 36, 16, 34, 44], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = f"A{hr + 1}"
    ws.sheet_view.showGridLines = False


def main() -> None:
    data = load_collections()
    settings = json.load(open(SETTINGS, encoding="utf-8"))
    pages = [json.loads(l) for l in open(PAGES, encoding="utf-8") if l.strip()]
    live = [p for p in pages if not page_is_junk(p)]
    redirects = json.load(open(REDIRECTS, encoding="utf-8")) if os.path.exists(REDIRECTS) else {"redirects": []}
    code_seo = json.load(open(CODE_SEO, encoding="utf-8")) if os.path.exists(CODE_SEO) else {}
    head_audit = json.load(open(HEAD_AUDIT, encoding="utf-8")) if os.path.exists(HEAD_AUDIT) else {}
    broken = json.load(open(BROKEN, encoding="utf-8")) if os.path.exists(BROKEN) else {}
    parity = json.load(open(PARITY, encoding="utf-8")) if os.path.exists(PARITY) else {}
    sm_parity = json.load(open(SITEMAP_PARITY, encoding="utf-8")) if os.path.exists(SITEMAP_PARITY) else {}
    og_cov = json.load(open(OG_COVERAGE, encoding="utf-8")) if os.path.exists(OG_COVERAGE) else {}
    legal = json.load(open(LEGAL, encoding="utf-8")) if os.path.exists(LEGAL) else {}

    wb = Workbook()
    build_summary(wb.active, settings, data, live, redirects,
                  {"parity": parity, "sitemap": sm_parity, "broken": broken})
    wb.active.title = "Summary"
    build_redirects(wb.create_sheet("Redirects"), redirects)
    build_url_parity(wb.create_sheet("URL Parity"), parity)
    build_sitemap_parity(wb.create_sheet("Sitemap Parity"), sm_parity)
    build_code_seo(wb.create_sheet("Global Code & SEO"), code_seo)
    build_head_audit(wb.create_sheet("Canonical & Head Audit"), head_audit)
    build_og_coverage(wb.create_sheet("OG Image Coverage"), og_cov)
    build_broken_links(wb.create_sheet("Broken Links"), broken)
    build_static_pages(wb.create_sheet("Static Pages"), live)
    build_legal(wb.create_sheet("Legal"), legal)
    new_desc = load_new_descriptions()
    for stem, (name, _) in COLLECTIONS.items():
        build_collection(wb.create_sheet(name), stem, data[stem], new_desc)
    wb.save(OUT)
    print(f"[xlsx] wrote {OUT}")
    print(f"  Sheets: Summary, Redirects, Global Code & SEO, Canonical & Head Audit, "
          f"Broken Links ({broken.get('counts', {}).get('broken', 0)}), "
          f"Static Pages ({len(live)}), {len(COLLECTIONS)} collections")


if __name__ == "__main__":
    main()
