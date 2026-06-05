#!/usr/bin/env python3
"""
Build the Webflow meta title/description Excel deliverable.

Reads migrations/webflow-export/seo-extract.csv (fresh Webflow source, which now
matches both the local and prod CMS DBs) and writes one sheet per collection plus
a Summary sheet. Blank/null meta is rendered as '-'.

Run: python3 migrations/webflow-import/build-seo-excel.py
"""
from __future__ import annotations

import csv
import os
from collections import OrderedDict

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

BASE = os.path.join(os.path.dirname(__file__), "..", "webflow-export")
SRC = os.path.join(BASE, "seo-extract.csv")
OUT = os.path.join(BASE, "webflow-meta-titles-descriptions.xlsx")

# collection stem -> (Sheet display name, public route prefix or None)
COLLECTIONS: "OrderedDict[str, tuple[str, str | None]]" = OrderedDict([
    ("blogs", ("Blogs", "/blogs")),
    ("news", ("News", "/news")),
    ("guides", ("Guides", "/guide")),
    ("events", ("Events", "/event")),
    ("jobs", ("Jobs", "/job")),
    ("authors", ("Authors", "/author")),
    ("resources", ("Resources", "/resources")),
    ("webinars", ("Webinars", "/webinar")),
    ("categories", ("Categories", "/category")),
    ("newsCategories", ("News Categories", "/news-categories")),
    ("jobLocations", ("Job Locations", None)),
    ("aboutGalleries", ("About Galleries", None)),
])

DASH = "-"
HEADERS = ["Name", "Slug", "URL", "Meta Title", "Meta Description"]
FONT = "Arial"

HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(name=FONT, bold=True, color="FFFFFF", size=11)
CELL_FONT = Font(name=FONT, size=10)
DASH_FONT = Font(name=FONT, size=10, color="B0B0B0")
TOP = Alignment(horizontal="left", vertical="top", wrap_text=True)
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WIDTHS = [34, 44, 46, 58, 82]


def load() -> "OrderedDict[str, list[dict[str, str]]]":
    data: "OrderedDict[str, list[dict[str, str]]]" = OrderedDict(
        (stem, []) for stem in COLLECTIONS
    )
    with open(SRC, encoding="utf-8") as fh:
        for row in csv.DictReader(fh):
            stem = row["collection"]
            if stem in data:
                data[stem].append(row)
    for rows in data.values():
        rows.sort(key=lambda r: r["slug"])
    return data


def cell_value(raw: str) -> str:
    raw = (raw or "").strip()
    return raw if raw else DASH


def write_collection_sheet(ws, stem: str, rows: list[dict[str, str]]) -> None:
    _, prefix = COLLECTIONS[stem]
    ws.append(HEADERS)
    for c in range(1, len(HEADERS) + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = BORDER

    for r in rows:
        slug = r["slug"].strip()
        url = f"{prefix}/{slug}" if (prefix and slug) else DASH
        ws.append([
            cell_value(r["name"]),
            cell_value(slug),
            url,
            cell_value(r["meta_title"]),
            cell_value(r["meta_description"]),
        ])

    for row_cells in ws.iter_rows(min_row=2, max_row=ws.max_row,
                                  min_col=1, max_col=len(HEADERS)):
        for cell in row_cells:
            cell.font = DASH_FONT if cell.value == DASH else CELL_FONT
            cell.alignment = TOP
            cell.border = BORDER

    for i, w in enumerate(WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False


def write_summary(ws, data) -> None:
    title = ws.cell(row=1, column=1, value="Webflow Meta Titles & Descriptions")
    title.font = Font(name=FONT, bold=True, size=14)
    sub = ws.cell(row=2, column=1,
                  value="Source: fresh Webflow CMS API pull (2026-06-05). "
                        "Values imported verbatim into local + prod CMS "
                        "(seo_title / seo_description). '-' = null/blank in Webflow.")
    sub.font = Font(name=FONT, italic=True, size=9, color="595959")

    head = ["Collection", "Total Items", "Has Meta Title", "Has Meta Description"]
    ws.append([])  # row 3 spacer
    ws.append(head)  # row 4
    hr = 4
    for c in range(1, len(head) + 1):
        cell = ws.cell(row=hr, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = BORDER

    row = hr + 1
    tot_n = tot_t = tot_d = 0
    for stem, (name, _) in COLLECTIONS.items():
        rows = data[stem]
        n = len(rows)
        has_t = sum(1 for r in rows if (r["meta_title"] or "").strip())
        has_d = sum(1 for r in rows if (r["meta_description"] or "").strip())
        tot_n += n
        tot_t += has_t
        tot_d += has_d
        ws.cell(row=row, column=1, value=name).font = CELL_FONT
        ws.cell(row=row, column=2, value=n).font = CELL_FONT
        ws.cell(row=row, column=3, value=has_t).font = CELL_FONT
        ws.cell(row=row, column=4, value=has_d).font = CELL_FONT
        for c in range(1, 5):
            ws.cell(row=row, column=c).border = BORDER
            ws.cell(row=row, column=c).alignment = Alignment(
                horizontal="left", vertical="center")
        row += 1

    tr = row
    bold = Font(name=FONT, bold=True, size=10)
    for c, val in enumerate(["TOTAL", tot_n, tot_t, tot_d], start=1):
        cell = ws.cell(row=tr, column=c, value=val)
        cell.font = bold
        cell.border = BORDER

    for i, w in enumerate([22, 14, 16, 20], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.sheet_view.showGridLines = False


def main() -> None:
    data = load()
    wb = Workbook()
    summary = wb.active
    summary.title = "Summary"
    write_summary(summary, data)
    for stem, (name, _) in COLLECTIONS.items():
        ws = wb.create_sheet(title=name)
        write_collection_sheet(ws, stem, data[stem])
    wb.save(OUT)
    print(f"[xlsx] wrote {OUT}")
    for stem, (name, _) in COLLECTIONS.items():
        print(f"  {name:18} {len(data[stem]):4d} rows")


if __name__ == "__main__":
    main()
